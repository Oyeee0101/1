import openpyxl

flag = 1
while flag:
    filename = input("plz enter the filename:")

    if filename == "exit":
        break
    else:
        flag = 1

    wb = openpyxl.load_workbook(filename + ".xlsx")
    ls = wb.sheetnames
    ws = wb[ls[0]]

    ls1 = ['L(t)','tL(t)','t^2L(t)','ΣL(t)','ΣtL(t)','Σt^2L(t)','期望','方差','无因次','n']
    for i in range(3,13,1):
        ws.cell(1,i).value = ls1[i-3]

    k = ws.max_row
    ls2 = [[],[],[]]   #L(t)  tL(t)  t^2L(t)

    for i in range(2,k+1):
        ws.cell(i,3).value = ws.cell(i,2).value -ws.cell(k,2).value
        ws.cell(i,4).value = ws.cell(i,3).value * ws.cell(i,1).value
        ws.cell(i,5).value = ws.cell(i,3).value * pow(ws.cell(i,1).value,2)
        for j in range(0,3):
            ls2[j].append(ws.cell(i,j+3).value)

    for i in range(3):
        ws.cell(2,i+6).value = sum(ls2[i])

    ws.cell(2,9).value = ws.cell(2,7).value / ws.cell(2,6).value
    ws.cell(2,10).value = ws.cell(2,8).value / ws.cell(2,6).value - pow(ws.cell(2,9).value,2)
    ws.cell(2,11).value = ws.cell(2,10).value / pow(ws.cell(2,9).value,2)
    ws.cell(2,12).value = 1/ws.cell(2,11).value

    wb.save(filename + '-calculated.xlsx')
